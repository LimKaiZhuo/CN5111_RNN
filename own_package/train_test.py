import keras.backend as K
import tensorflow as tf
import gc
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import time, os
# Own Scripts
from own_package.models import LSTMmodel
from .others import print_array_to_excel
from .features_labels_setup import load_data_to_fl
from own_package.others import create_results_directory


def run_train_test(model_mode, hparams, window_size, loader_file, results_directory=None, seed=42,
                   save_model=False, save_model_name=None):
    '''
    Stratified k fold cross validation for training and evaluating model 2 only. Model 1 data is trained before hand.
    :param model_mode: Choose between using SNN or cDNN (non_smiles) and SNN_smiles or cDNN_smiles
    :param cv_mode: Cross validation mode. Either 'skf' or 'loocv'.
    :param hparams: hparams dict containing hyperparameters information
    :param loader_file: data_loader excel file location
    :param skf_file: skf_file name to save excel file as
    :param skf_sheet: name of sheet to save inside the skf_file excel. If None, will default to SNN or cDNN as name
    :param k_folds: Number of k folds. Used only for skf cv_mode
    :param k_shuffle: Whether to shuffle the given examples to split into k folds if using skf
    :return:
    '''
    if not results_directory:
        results_directory = './results/{}'.format(model_mode)
    results_directory = create_results_directory(results_directory)

    fl = load_data_to_fl(loader_file, window_size=window_size)

    # Run train test
    sess = tf.Session()
    K.set_session(sess)
    instance_start = time.time()
    (ss_fl, i_ss_fl) = fl.create_train_test_split(seed=seed)  # ss_fl is training fl, i_ss_fl is validation fl

    # Set up model
    model = LSTMmodel(ss_fl, model_mode, hparams)
    # Train model and save model training loss vs epoch plot if plot_name is given, else no plot will be saved
    model.train_model(ss_fl, save_mode=False, plot_name='{}/plots/training_loss.png'.format(results_directory))

    # Evaluation
    predicted_labels, mse = model.eval(i_ss_fl)

    # Saving model
    if save_model:
        # Set save_model_name
        if isinstance(save_model_name, str):
            save_model_name1 = save_model_name + '_' + model_mode
        else:
            save_model_name1 = model_mode
        # Save model
        dirc = results_directory + '/models/' + save_model_name1 + '.h5'
        print('Saving model in {}'.format(dirc))
        model.model.save(dirc)

    # Need to put the next 3 lines if not memory will run out
    del model
    K.clear_session()
    gc.collect()

    # Printing one instance summary.
    instance_end = time.time()

    print('Model is {}. Time take for instance = {}\n'
          'Post-training results: \nmse = {},\n'
          '####################################################################################################'
          .format(model_mode, instance_end - instance_start, mse))

    # Plotting the time series plot for prediction and actual test labels
    for k in range(i_ss_fl.count):
        plt.plot(np.squeeze(i_ss_fl.labels[k, :, 0]), c='g', label='Actual')
        plt.plot(np.squeeze(predicted_labels[k, :]), c='r', label='Predicted')
        plt.legend(loc='upper left')
        plt.title('Test Example ' + str(k + 1))
        plt.ylabel('Demand')
        plt.xlabel('Hours of the day')
        plt.savefig(results_directory + '/plots/validation_plots/Test Example ' + str(k + 1) + '.png',
                    bbox_inches='tight')
        plt.close()

    # Printing results to excel
    # Creating excel
    excel_name = results_directory + '/results.xlsx'
    wb = openpyxl.Workbook()
    wb.save(excel_name)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]

    # Writing other subset split, instance per run, and bounds
    start_row = 1
    start_col = 1
    headers = ['mse']
    values = [mse]
    print_array_to_excel(np.array(headers), (start_row, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values), (1+start_row, start_col + 1), ws, axis=1)
    start_col +=2

    # Writing hparams dataframe
    pd_writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    hparams = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in hparams.items()]))
    hparams.to_excel(pd_writer, sheetname, startrow=0, startcol=start_col)

    # Saving and closing
    pd_writer.save()
    pd_writer.close()
    wb.close()

    return mse
