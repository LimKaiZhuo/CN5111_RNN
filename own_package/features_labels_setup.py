import numpy as np
import numpy.random as rng
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split
from keras.utils import to_categorical
import pickle
import os
import pathlib
import warnings
import copy
import xlrd
from .others import print_array_to_excel


def flat_to_time(data, window_size):
    """
    Input flatten features_c and labels
    :return: Reshaped according to [batches, window_size, number of features/ labels]
    """
    assert len(data.shape) == 2, 'data shape is not 2. [flatten sample, features/ labels dim]'
    return data.reshape(-1, window_size, data.shape[1])


def time_to_flat(data):
    """
    Input time features_c and labels
    :return: Reshaped according to [flatten sample, number of features/ labels]
    """
    assert len(data.shape) == 3, 'data shape is not 3. [sample, time idx, features/ labels dim]'
    return data.reshape(-1, data.shape[2])


def load_data_to_fl(data_loader_excel_file, window_size=24):
    df = pd.read_excel(data_loader_excel_file, sheet_name='raw')

    features_c = df.iloc[:, 0].values
    features_c_names = df.columns.values[0]

    labels = df.iloc[:, 1].values
    labels_names = df.columns.values[1]

    features_c = np.reshape(features_c, (-1, 1))
    labels = np.reshape(labels, (-1, 1))

    features_c = flat_to_time(features_c, window_size)
    labels = flat_to_time(labels, window_size)

    fl = Features_labels(features_c, labels, features_c_names, labels_names, window_size, scaler=None)

    return fl


class Features_labels:
    def __init__(self, features_c, labels, features_c_names, labels_names, window_size, scaler=None):
        """
        Creates fl class with a lot useful attributes
        :param features_c: Continuous features. Np array, no. of examples x continous features
        :param labels: Labels as np array, no. of examples x dim
        :param scaler: Scaler to transform features c. If given, use given MinMax scaler from sklearn,
        else create scaler based on given features c.
        """
        assert len(features_c.shape) == 3, 'features_c shape is not 3. [sample, time idx, features]'
        assert len(labels.shape) == 3, 'labels shape is not 3. [sample, time idx, labels=1]'

        self.window_size = window_size
        # Setting up features
        self.count = features_c.shape[0]
        self.features_c_names = features_c_names
        self.features_c_count = features_c.shape[1]
        self.features_c = np.copy(features_c)
        self.features_c_dim = features_c.shape[1]
        self.features_c_f = time_to_flat(features_c)

        # Scaling features_c
        if scaler is None:
            # If scaler is None, means normalize the data with all input data
            self.scaler = MinMaxScaler()
            self.scaler.fit(self.features_c_f)  # Setting up scaler
        else:
            # If scaler is given, means normalize the data with the given scaler
            self.scaler = scaler

        self.features_c_norm_f = self.scaler.transform(self.features_c_f)  # Normalizing features_c
        self.features_c_norm = flat_to_time(self.features_c_norm_f, self.window_size)

        # Setting up labels
        self.labels = labels
        self.labels_names = labels_names

    def create_train_test_split(self, seed=42):
        x_train, x_test, y_train, y_test = train_test_split(self.features_c, self.labels, random_state=seed,
                                                            test_size=0.2)
        return [Features_labels(features_c=x_train, labels=y_train,
                               features_c_names=self.features_c_names, labels_names=self.labels_names,
                               window_size=self.window_size, scaler=self.scaler), \
               Features_labels(features_c=x_test,
                               labels=y_test,
                               features_c_names=self.features_c_names,
                               labels_names=self.labels_names,
                               window_size=self.window_size,
                               scaler=self.scaler)]

    def write_data_to_excel(self, loader_excel_file='./excel/data_loader.xlsx'):
        # Excel writing part
        wb = load_workbook(loader_excel_file)
        # Setting up subset features and labels sheet
        sheet_name_store = ['temp_features_c', 'temp_labels']
        ss_store = [self.features_c, self.labels]
        axis_store = [2, 0]  # Because feature_c is 2D while labels is col vector, so order of axis is 2,0
        for cnt, sheet_name in enumerate(sheet_name_store):
            if sheet_name in wb.sheetnames:
                # If temp sheet exists, remove it to create a new one. Else skip this.
                idx = wb.sheetnames.index(sheet_name)  # index of temp sheet
                wb.remove(wb.worksheets[idx])  # remove temp
                wb.create_sheet(sheet_name, idx)  # create an empty sheet using old index
            else:
                wb.create_sheet(sheet_name)  # Create the new sheet
            # Print array to the correct worksheet
            print_array_to_excel(ss_store[cnt], (2, 1), wb[sheet_name_store[cnt]], axis=axis_store[cnt])
        wb.save(loader_excel_file)
        wb.close()
